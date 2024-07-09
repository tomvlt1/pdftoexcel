import os
import re
import fitz  # PyMuPDF
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo

def extract_table_from_pdf(pdf_path):
    """
    Extracts table data from a PDF file.

    Args:
        pdf_path (str): Path to the PDF file.

    Outputs:
        - List of strings, each representing a line of the extracted table data.
        - Returns an empty list if there is an error opening the PDF or reading any page.

    Returns:
        list: A list of strings representing the extracted table data.
    """

    try:
        pdf_document = fitz.open(pdf_path)
    except Exception as e:
        print(f"Error opening PDF {pdf_path}: {e}")
        return []
    
    all_data = []
    exclude_phrases = ["FIN DE LA LISTE", "**************************", "-----------------------"] # fin de chaque token, eliminer de la liste

    for page_num in range(pdf_document.page_count):
        try:
            page = pdf_document.load_page(page_num)
            text = page.get_text("text")
        except Exception as e:
            print(f"Error reading page {page_num} of {pdf_path}: {e}")
            continue

        lines = text.split('\n')
        table_started = False
        page_data = []
        for line in lines:
            if table_started:
                if line.strip() == '' or any(phrase in line for phrase in exclude_phrases):
                    continue
                else:
                    page_data.append(line)
            elif '--------------------' in line: # also determine the beginning of a section
                table_started = True

        if page_data:
            all_data.extend(page_data)

    return all_data

def split_line(line):
    """
    Splits a line of text into components based on whitespace and cleans the data.

    Args:
        line (str): A single line of text.

    Outputs:
        - List of strings, each representing a part of the split line.
        - The first part is the number extracted from the beginning of the line.
        - The last part is cleaned of any trailing question marks.

    Returns:
        list: A list of strings representing the split line components.
    """

    split_data = re.split(r'\s{3,}', line)

    if split_data:
        first_element = split_data[0]
        number = first_element.split(' ', 1)[0]  # So that only the number is extracted
        split_data = [number] + split_data[1:]

    if split_data and len(split_data) > 1:
        split_data[-1] = re.sub(r'\?+', '', split_data[-1])
    
    return split_data

def create_excel_table(ws, table_name, ref):
    """
    Creates an Excel table within a worksheet.

    Args:
        ws (openpyxl.worksheet.worksheet.Worksheet): The worksheet to add the table to.
        table_name (str): The name of the table.
        ref (str): The reference range for the table.

    Outputs:
        - An Excel table added to the specified worksheet.

    Returns:
        None
    """


    tab = Table(displayName=table_name, ref=ref)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab.tableStyleInfo = style
    ws.add_table(tab)

def process_files_in_folder(input_folder, output_excel_path):
    """
    Processes PDF and Excel files in a specified folder and combines data into an output Excel file.

    Args:
        input_folder (str): Path to the input folder containing PDFs and Excel files.
        output_excel_path (str): Path to the output Excel file.

    Outputs:
        - An Excel file at the specified output path containing combined data.
        - A 'Combined Data' sheet with data from all processed PDFs.
        - A 'Summary' sheet with specified headers and formulas.
        - Sheets for each unique 'instrument_id' with filtered data.

    Returns:
        None
    """

    if not os.path.exists(input_folder):
        print(f"Input folder {input_folder} does not exist.")
        return
    
    combined_data = []
    wb = Workbook()
    ws_combined = wb.active
    ws_combined.title = "Combined Data"
    
    instrument_id_df = None
    import_tdx_df = None

    for filename in os.listdir(input_folder):
        file_path = os.path.join(input_folder, filename)
        
        if not os.path.isfile(file_path):
            continue
        
        if filename.endswith(".pdf"):
            table_data = extract_table_from_pdf(file_path)
            if table_data:
                combined_data.append([filename]) 
                for line in table_data:
                    combined_data.append([line])  
                    split_data = split_line(line)
                    combined_data.append([''] * 9 + split_data)  
                combined_data.append([''] * 16)  
                combined_data.append([''] * 16)

        elif filename.endswith(".xlsx"):
            print(f"Processing Excel file: {filename}")
            try:
               
                df = pd.read_excel(file_path)
                sheet_name = os.path.splitext(filename)[0]
                ws = wb.create_sheet(title=sheet_name)
                for row in dataframe_to_rows(df, index=False, header=True):
                    ws.append(row)

                if "instrument_id" in sheet_name.lower(): #check if the sheet is actually within the excel - NAME CORRECTLY
                    instrument_id_df = df
                    print(f"Loaded instrument_ids sheet from {filename}")

                if "import tdx" in sheet_name.lower(): #check if the sheet is actually within the excel - NAME CORRECTLY
                    import_tdx_df = df
                    print(f"Loaded import tdx sheet from {filename}")

            except Exception as e:
                print(f"Error reading Excel file {file_path}: {e}")

    if combined_data:
        df_combined = pd.DataFrame(combined_data)
        for row in dataframe_to_rows(df_combined, index=False, header=False):
            ws_combined.append(row)

    
    if instrument_id_df is not None and import_tdx_df is not None: # check if both dataframes have been loaded
        print("Both instrument_id and import_tdx dataframes are loaded.")
        for _, row in instrument_id_df.iterrows():
            try:
                instrument_id = row['instrument id'] # both of these are name and case sensitive make sure they match exactly 
                best_name = row['best name']
            except KeyError as e:
                print("Incorrect column name")
                print(f"Available columns in instrument_id_df: {instrument_id_df.columns}") # allows for debugging in case of errors, tells the user what to name their columns
                continue
            
            filtered_df = import_tdx_df[import_tdx_df['InstrumentID'] == instrument_id]

            if not filtered_df.empty:
                sheet_name = best_name[:31] 
                ws = wb.create_sheet(title=sheet_name)
                for df_row in dataframe_to_rows(filtered_df, index=False, header=True):
                    ws.append(df_row)
                create_excel_table(ws, table_name=sheet_name, ref=f"A1:{chr(64+len(filtered_df.columns))}{len(filtered_df)+1}")
                print(f"Created sheet {sheet_name} for InstrumentID {instrument_id}")
    else:
        print("Missing one or both dataframes: instrument_id and import_tdx. Make sure they are correctly named AND they are the correct file type, not csv but excel")

    ws_summary = wb.create_sheet(title="Summary")
    headers = ["id client", "best_id_imported", "balance best", "token", "balance tdx", "Diff"]
    ws_summary.append(headers)

   
    formulas = [  # paste the formulas as text in row 3 of the final sheet
        '=IF(RIGHT([@token],3)="pdf","fichier",VLOOKUP(B2, \'id vs best\'!$A$2:$B$3000, 2, FALSE))',
        '=IFERROR(VALUE(TRIM(CLEAN(LEFT(SUBSTITUTE(\'Combined Data\'!K2, ".", ""), IFERROR(FIND(" ", SUBSTITUTE(\'Combined Data\'!K2, ".", "")) - 1, LEN(SUBSTITUTE(\'Combined Data\'!K2, ".", ""))))))), "")',
        '=IFERROR(VALUE(TRIM(CLEAN(SUBSTITUTE(SUBSTITUTE(IF(\'Combined Data\'!M2="","",\'Combined Data\'!M2), "\'", ""), ",", ".")))),"")',
        '=CLEAN(TRIM(IF(RIGHT(\'Combined Data\'!A1, 3) = "pdf", \'Combined Data\'!A1, IFERROR(LEFT(D1, FIND(".pdf", D1) - 1), D1))))',
        '=VLOOKUP([@[id client]], INDIRECT([@[token]] & "[#All]"), 9, FALSE)',
        '=IFERROR([@[balance best]]-[@[balance tdx]]," ")'
    ]
    ws_summary.append([""] * len(headers))  # add an empty row before the formulas
    ws_summary.append([f'"{formula}"' for formula in formulas])  # add formulas in row 3
    ws_summary.append(["Paste the formula's in the cells above and then delete this line"])
    ws_summary.append(["drag down the formula until where you need them"])
    ws_summary.append(["in ID client filter out the NA's and in best_id filter out the 1 id AFTER you have dragged the formulas all the way down"])
    create_excel_table(ws_summary, table_name="SummaryTable", ref="A1:F2")

    try:
        wb.save(output_excel_path)
        print(f"Combined table and Excel sheets saved to {output_excel_path}")
    except Exception as e:
        print(f"Error saving Excel file {output_excel_path}: {e}")

input_folder = "input"  # change this to the path of your folder containing PDFs and Excel files
output_excel_path = "output/combined_output.xlsx"  # change this to the path of the combined Excel file

process_files_in_folder(input_folder, output_excel_path)
